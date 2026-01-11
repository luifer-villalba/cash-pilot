# File: src/cashpilot/api/export_sessions.py
"""Export endpoints for cash sessions data (CSV/Excel)."""

import csv
import io
from datetime import date as date_type
from decimal import Decimal
from typing import Literal

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from cashpilot.api.auth_helpers import require_admin
from cashpilot.core.db import get_db
from cashpilot.core.logging import get_logger
from cashpilot.models import CashSession, User
from cashpilot.models.user import UserRole

logger = get_logger(__name__)

router = APIRouter(prefix="/api/export", tags=["export"])


def _format_paraguayan_number(value: Decimal | None) -> str:
    """Format decimal as Paraguayan number format (dots for thousands, comma for decimals).

    Example: 1500000.50 -> 1.500.000,50
    """
    if value is None:
        return ""

    # Convert to string with 2 decimal places
    str_value = f"{value:.2f}"
    integer_part, decimal_part = str_value.split(".")

    # Add dots for thousands separator
    formatted_int = ""
    for i, digit in enumerate(reversed(integer_part)):
        if i > 0 and i % 3 == 0:
            formatted_int = "." + formatted_int
        formatted_int = digit + formatted_int

    # Use comma for decimal separator
    return f"{formatted_int},{decimal_part}"


def _format_boolean(value: bool) -> str:
    """Format boolean for export."""
    return "Yes" if value else "No"


async def _get_sessions_for_export(
    db: AsyncSession,
    current_user: User,
    from_date: str | None = None,
    to_date: str | None = None,
    cashier_name: str | None = None,
    business_id: str | None = None,
    status: str | None = None,
) -> list[CashSession]:
    """Fetch sessions for export with filters."""

    stmt = select(CashSession).where(~CashSession.is_deleted)

    # Apply role-based filtering
    if current_user.role == UserRole.CASHIER:
        stmt = stmt.where(CashSession.cashier_id == current_user.id)

    # Apply date range filters
    if from_date:
        try:
            from_date_obj = date_type.fromisoformat(from_date)
            stmt = stmt.where(CashSession.session_date >= from_date_obj)
        except (ValueError, TypeError):
            pass  # Invalid date, skip filter

    if to_date:
        try:
            to_date_obj = date_type.fromisoformat(to_date)
            stmt = stmt.where(CashSession.session_date <= to_date_obj)
        except (ValueError, TypeError):
            pass  # Invalid date, skip filter

    # Apply business filter
    if business_id:
        try:
            from uuid import UUID

            business_uuid = UUID(business_id)
            stmt = stmt.where(CashSession.business_id == business_uuid)
        except (ValueError, TypeError):
            pass  # Invalid UUID, skip filter

    # Apply status filter
    if status and status.upper() in ["OPEN", "CLOSED"]:
        stmt = stmt.where(CashSession.status == status.upper())

    # Apply cashier name filter (searches in cashier first_name/last_name/display_name)
    if cashier_name:
        from cashpilot.models import User as UserModel

        stmt = stmt.join(UserModel, CashSession.cashier_id == UserModel.id)
        ilike = f"%{cashier_name}%"
        stmt = stmt.where(
            (UserModel.first_name.ilike(ilike))
            | (UserModel.last_name.ilike(ilike))
            | (UserModel.email.ilike(ilike))
        )

    # Order by date descending (most recent first)
    stmt = stmt.order_by(CashSession.session_date.desc(), CashSession.opened_time.desc())

    result = await db.execute(stmt)
    return list(result.scalars().all())


def _get_export_headers() -> list[str]:
    """Get column headers for export."""
    return [
        "Session ID",
        "Session Number",
        "Date",
        "Business Name",
        "Cashier Name",
        "Status",
        "Opened Time",
        "Closed Time",
        "Initial Cash",
        "Final Cash",
        "Cash Sales",
        "Card Total",
        "Bank Transfer",
        "Credit Sales",
        "Credit Collected",
        "Total Sales",
        "Expenses",
        "Net Earnings",
        "Envelope Amount",
        "Discrepancy",
        "Flagged",
        "Flag Reason",
        "Notes",
        "Closing Ticket",
    ]


def _session_to_row(session: CashSession, format_type: Literal["csv", "excel"]) -> list:
    """Convert a CashSession to a row of data.

    Args:
        session: The cash session to convert
        format_type: 'csv' or 'excel' - determines number formatting
    """
    # Calculate discrepancy
    if session.final_cash is not None:
        expected = session.initial_cash + session.cash_sales - session.envelope_amount
        discrepancy = session.final_cash - expected
    else:
        discrepancy = None

    # For CSV, use Paraguayan format; for Excel, use raw numbers
    if format_type == "csv":
        format_num = _format_paraguayan_number
    else:
        # For Excel, return raw Decimal/float for proper cell formatting
        def format_num(x):
            return float(x) if x is not None else None

    return [
        str(session.id),
        session.session_number,
        session.session_date.isoformat(),
        session.business.name if session.business else "",
        session.cashier.display_name if session.cashier else "",
        session.status,
        session.opened_time.strftime("%H:%M:%S") if session.opened_time else "",
        session.closed_time.strftime("%H:%M:%S") if session.closed_time else "",
        format_num(session.initial_cash),
        format_num(session.final_cash),
        format_num(session.cash_sales),
        format_num(session.card_total),
        format_num(session.bank_transfer_total),
        format_num(session.credit_sales_total),
        format_num(session.credit_payments_collected),
        format_num(session.total_sales),
        format_num(session.expenses),
        format_num(session.net_earnings),
        format_num(session.envelope_amount),
        format_num(discrepancy),
        _format_boolean(session.flagged),
        session.flag_reason or "",
        session.notes or "",
        session.closing_ticket or "",
    ]


def _create_csv_export(sessions: list[CashSession]) -> str:
    """Generate CSV content from sessions."""
    output = io.StringIO()
    writer = csv.writer(output)

    # Write headers
    writer.writerow(_get_export_headers())

    # Write data rows
    for session in sessions:
        writer.writerow(_session_to_row(session, format_type="csv"))

    return output.getvalue()


def _create_excel_export(sessions: list[CashSession]) -> bytes:
    """Generate Excel file from sessions with Paraguayan number formatting."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Cash Sessions"

    # Write headers with styling
    headers = _get_export_headers()
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Write data rows
    for row_idx, session in enumerate(sessions, start=2):
        row_data = _session_to_row(session, format_type="excel")
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            # Apply Paraguayan number format to currency columns
            # Format: #.##0,00 (dots for thousands, comma for decimal)
            currency_columns = [
                9,
                10,
                11,
                12,
                13,
                14,
                15,
                16,
                17,
                18,
                19,
                20,
            ]  # Initial Cash through Discrepancy
            if col_idx in currency_columns and isinstance(value, (int, float, Decimal)):
                cell.number_format = "#,##0.00"
                # Note: Excel uses regional settings for separator display
                # We'll add a custom format that explicitly uses dots and commas
                cell.number_format = "#.##0,00"

    # Auto-adjust column widths
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)

        # Set minimum width based on header
        max_length = len(header)

        # Adjust for data (sample first 100 rows for performance)
        for row_idx in range(2, min(102, len(sessions) + 2)):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))

        # Set column width with some padding
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


@router.get("/sessions")
async def export_sessions(
    format: Literal["csv", "xlsx"] = Query("csv", description="Export format: csv or xlsx"),
    from_date: str | None = Query(None, description="Start date (YYYY-MM-DD)"),
    to_date: str | None = Query(None, description="End date (YYYY-MM-DD)"),
    cashier_name: str | None = Query(None, description="Filter by cashier name"),
    business_id: str | None = Query(None, description="Filter by business ID"),
    status: str | None = Query(None, description="Filter by status (OPEN/CLOSED)"),
    current_user: User = Depends(require_admin),
    db: AsyncSession = Depends(get_db),
):
    """Export cash sessions data to CSV or Excel.

    **Admin only** - Exports cash sessions with applied filters.

    Features:
    - CSV format with Paraguayan number formatting (1.500.000,50)
    - Excel format with proper cell formatting and styling
    - Date range filtering
    - Business and cashier filtering
    - Status filtering

    Query Parameters:
    - `format`: csv or xlsx (default: csv)
    - `from_date`: Start date in YYYY-MM-DD format
    - `to_date`: End date in YYYY-MM-DD format
    - `cashier_name`: Filter by cashier name (partial match)
    - `business_id`: Filter by business UUID
    - `status`: Filter by status (OPEN or CLOSED)
    """

    # Fetch sessions with filters
    sessions = await _get_sessions_for_export(
        db=db,
        current_user=current_user,
        from_date=from_date,
        to_date=to_date,
        cashier_name=cashier_name,
        business_id=business_id,
        status=status,
    )

    logger.info(
        "export.sessions",
        format=format,
        count=len(sessions),
        user_id=str(current_user.id),
        filters={
            "from_date": from_date,
            "to_date": to_date,
            "cashier_name": cashier_name,
            "business_id": business_id,
            "status": status,
        },
    )

    # Generate export file
    from datetime import datetime

    now_str = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    if format == "xlsx":
        content = _create_excel_export(sessions)
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename = f"cash_sessions_export_{now_str}.xlsx"
    else:  # csv
        content = _create_csv_export(sessions)
        media_type = "text/csv; charset=utf-8"
        filename = f"cash_sessions_export_{now_str}.csv"

    # Return as downloadable file
    return StreamingResponse(
        io.BytesIO(content.encode("utf-8") if isinstance(content, str) else content),
        media_type=media_type,
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Cache-Control": "no-cache, no-store, must-revalidate",
            "Pragma": "no-cache",
            "Expires": "0",
        },
    )
