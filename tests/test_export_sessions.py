# File: tests/test_export_sessions.py
"""Tests for session export endpoints."""

import csv
import io
from datetime import date, time
from decimal import Decimal

import pytest
from httpx import AsyncClient
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from tests.factories import BusinessFactory, CashSessionFactory, UserFactory


@pytest.fixture
async def admin_user(db_session: AsyncSession):
    """Create an admin user for testing."""
    from cashpilot.models.user import UserRole
    
    user = await UserFactory.create(
        db_session,
        email="admin_export@test.com",
        full_name="Export Admin",
        role=UserRole.ADMIN,
        is_active=True,
    )
    return user


@pytest.fixture
async def admin_export_client(app, admin_user):
    """Create an authenticated admin client."""
    from httpx import AsyncClient
    
    async with AsyncClient(app=app, base_url="http://test") as client:
        # Set up session to simulate logged-in admin
        client.cookies.set("session", "test-admin-session")
        client.test_user = admin_user
        yield client


@pytest.fixture
async def sample_sessions(db_session: AsyncSession, admin_user):
    """Create sample cash sessions for export testing."""
    business = await BusinessFactory.create(
        db_session,
        name="Export Test Business",
        address="Test Address 123",
        phone="+595972123456",
    )
    
    cashier = await UserFactory.create(
        db_session,
        email="cashier_export@test.com",
        full_name="Test Cashier",
        role="CASHIER",
        is_active=True,
    )
    
    # Create multiple sessions with different data
    sessions = []
    for i in range(5):
        session = await CashSessionFactory.create(
            db_session,
            business_id=business.id,
            cashier_id=cashier.id,
            created_by=admin_user.id,
            session_date=date(2026, 1, 10),
            opened_time=time(8, 0),
            closed_time=time(16, 0) if i % 2 == 0 else None,
            status="CLOSED" if i % 2 == 0 else "OPEN",
            initial_cash=Decimal("100000.00"),
            final_cash=Decimal("150000.00") if i % 2 == 0 else None,
            card_total=Decimal("50000.00"),
            bank_transfer_total=Decimal("25000.00"),
            expenses=Decimal("10000.00"),
            flagged=i == 0,  # Flag first session
            flag_reason="Test flag reason" if i == 0 else None,
            notes=f"Session {i} notes",
        )
        sessions.append(session)
    
    await db_session.commit()
    return sessions


class TestExportSessions:
    """Test session export functionality."""
    
    @pytest.mark.asyncio
    async def test_export_csv_requires_admin(
        self, client: AsyncClient, sample_sessions
    ):
        """Test that CSV export requires admin role."""
        # Non-admin user should get 403
        response = await client.get("/api/export/sessions?format=csv")
        assert response.status_code == 403
    
    @pytest.mark.asyncio
    async def test_export_csv_success(
        self, admin_export_client: AsyncClient, sample_sessions
    ):
        """Test successful CSV export with admin user."""
        response = await admin_export_client.get("/api/export/sessions?format=csv")
        
        # Check response
        assert response.status_code == 200
        assert response.headers["content-type"].startswith("text/csv")
        assert "cash_sessions_export" in response.headers["content-disposition"]
        
        # Parse CSV content
        content = response.text
        csv_reader = csv.DictReader(io.StringIO(content))
        rows = list(csv_reader)
        
        # Verify we got sessions
        assert len(rows) >= 5
        
        # Verify headers
        expected_headers = [
            "Session ID", "Session Number", "Date", "Business Name", "Cashier Name",
            "Status", "Opened Time", "Closed Time", "Initial Cash", "Final Cash",
            "Cash Sales", "Card Total", "Bank Transfer", "Credit Sales",
            "Credit Collected", "Total Sales", "Expenses", "Net Earnings",
            "Envelope Amount", "Discrepancy", "Flagged", "Flag Reason", "Notes",
            "Closing Ticket"
        ]
        assert list(csv_reader.fieldnames) == expected_headers
        
        # Verify data format (Paraguayan format: dots for thousands, comma for decimal)
        first_row = rows[0]
        assert first_row["Business Name"] == "Export Test Business"
        assert first_row["Cashier Name"] == "Test Cashier"
        # Check Paraguayan number format (e.g., "100.000,00")
        assert "," in first_row["Initial Cash"]  # Has decimal comma
    
    @pytest.mark.asyncio
    async def test_export_xlsx_success(
        self, admin_export_client: AsyncClient, sample_sessions
    ):
        """Test successful Excel export with admin user."""
        response = await admin_export_client.get("/api/export/sessions?format=xlsx")
        
        # Check response
        assert response.status_code == 200
        assert "spreadsheet" in response.headers["content-type"]
        assert "cash_sessions_export" in response.headers["content-disposition"]
        assert response.headers["content-disposition"].endswith('.xlsx"')
        
        # Load Excel file
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Verify headers (row 1)
        headers = [cell.value for cell in ws[1]]
        assert "Session ID" in headers
        assert "Business Name" in headers
        assert "Total Sales" in headers
        
        # Verify we have data rows
        assert ws.max_row >= 6  # Header + at least 5 data rows
        
        # Verify data
        assert ws.cell(2, 4).value == "Export Test Business"  # Business Name in row 2
        assert ws.cell(2, 5).value == "Test Cashier"  # Cashier Name
    
    @pytest.mark.asyncio
    async def test_export_with_date_filter(
        self, admin_export_client: AsyncClient, sample_sessions
    ):
        """Test export with date range filter."""
        response = await admin_export_client.get(
            "/api/export/sessions?format=csv&from_date=2026-01-10&to_date=2026-01-10"
        )
        
        assert response.status_code == 200
        
        # Parse CSV
        content = response.text
        csv_reader = csv.DictReader(io.StringIO(content))
        rows = list(csv_reader)
        
        # All sessions should be from the filtered date
        for row in rows:
            assert row["Date"] == "2026-01-10"
    
    @pytest.mark.asyncio
    async def test_export_with_status_filter(
        self, admin_export_client: AsyncClient, sample_sessions
    ):
        """Test export with status filter."""
        response = await admin_export_client.get(
            "/api/export/sessions?format=csv&status=CLOSED"
        )
        
        assert response.status_code == 200
        
        # Parse CSV
        content = response.text
        csv_reader = csv.DictReader(io.StringIO(content))
        rows = list(csv_reader)
        
        # All sessions should be CLOSED
        for row in rows:
            assert row["Status"] == "CLOSED"
        
        # Should have only closed sessions (3 out of 5 in our fixture)
        assert len(rows) == 3
    
    @pytest.mark.asyncio
    async def test_export_includes_flagged_status(
        self, admin_export_client: AsyncClient, sample_sessions
    ):
        """Test that export includes flagged status and reason."""
        response = await admin_export_client.get("/api/export/sessions?format=csv")
        
        assert response.status_code == 200
        
        # Parse CSV
        content = response.text
        csv_reader = csv.DictReader(io.StringIO(content))
        rows = list(csv_reader)
        
        # Find the flagged session
        flagged_sessions = [r for r in rows if r["Flagged"] == "Yes"]
        assert len(flagged_sessions) >= 1
        
        # Verify flag reason is included
        flagged_session = flagged_sessions[0]
        assert flagged_session["Flag Reason"] == "Test flag reason"
